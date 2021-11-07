from flask import Flask, render_template, session, flash, redirect, request
from flask_app import app
from flask_bcrypt import Bcrypt
from flask_app.models.item import Item
from flask_app.models.warehouse import Warehouse
from flask_app.models.quantity import Quantity
from flask_app.models.planning import Planning
from flask_app.controllers.users import logged_in
from openpyxl import workbook, worksheet, load_workbook, writer


#  TODO module for being able to accept CSV

@app.route("/import")
@logged_in
def import_files():
    return render_template("import.html")


@app.route("/import/warehouses", methods=["POST"])
@logged_in
def import_warehouses():
    if request.method != "POST":
        return redirect("/")
    else:
        uploaded_file = request.files["warehouse_imports"]
        mappings = [request.form["warehouse_code"],
                    request.form["warehouse_name"],
                    request.form["warehouse_description"]]

        # spreadsheet columns will be:
        #   mappings[0] = Warehouse Code
        #   mappings[1] = Warehouse Name
        #   mappings[2] = Warehouse Description

        wb = load_workbook(uploaded_file)
        sheet = wb.active
        if request.form.get("ignore_first_row") == "on":
            start = 2
        else:
            start = 1
        end = sheet.max_row + 1

        for i in range(start, end):
            warehouse = {
                "code": sheet[f"{mappings[0]}{i}"].value,
                "warehouse_name": sheet[f"{mappings[1]}{i}"].value,
                "description": sheet[f"{mappings[2]}{i}"].value,
                "updated_by": session["user_id"]
            }
            for key in warehouse:
                if warehouse[key] is None:
                    warehouse[key] = ""
            log_row = warehouse["code"] + " " + warehouse["warehouse_name"]
            validation = Warehouse.validate_warehouse(warehouse, imported=True)
            if validation[0]:
                Warehouse.save_warehouse(warehouse)
                log_row += " Import successful!"
                flash(log_row, "import_successful")
            else:
                log_row += f" Import failed: {validation[1]}"
                flash(log_row, "import_fail")
            print(log_row)
        return redirect("/import/result")


@app.route("/import/items", methods=["POST"])
@logged_in
def import_items():
    if request.method != "POST":
        return redirect("/")
    else:
        uploaded_file = request.files["item_imports"]
        mappings = [request.form["item_number"],
                    request.form["item_description"]]
        # mappings[0] = item number
        # mappings[1] = item description

        wb = load_workbook(uploaded_file)
        sheet = wb.active
        end = sheet.max_row + 1
        if request.form.get("ignore_first_row") == "on":
            start = 2
        else:
            start = 1
        for i in range(start, end):
            item = {"item_number": sheet[f"{mappings[0]}{i}"].value, "description": sheet[f"{mappings[1]}{i}"].value}
            for key in item:
                if item[key] == None:
                    item[key] = ""
            log_row = item["item_number"] + " " + item["description"]
            validation = Item.validate_item(item, imported=True)
            if validation[0]:
                log_row += " Import successful!"
                flash(log_row, "import_successful")
                Item.save_item(item)
            else:
                for error in validation[1]:
                    log_row += (" " + error)
                flash(log_row, "import_fail")
        return redirect("/import/result")


@app.route("/import/warehouse_items", methods=["POST"])
@logged_in
def import_warehouse_items():
    mappings = [
        request.form["warehouse_item_number"],
        request.form["item_warehouse"],
        request.form["min"],
        request.form["max"],
        request.form["starting_qty"]
    ]

    file = request.files["warehouse_item_imports"]
    wb = load_workbook(file)
    sheet = wb.active
    if request.form.get("ignore_first_row") == "on":
        start = 2
    else:
        start = 1
    end = sheet.max_row + 1
    for i in range(start, end):

        data = {
            "item_number": sheet[f"{mappings[0]}{i}"].value,
            "code": sheet[f"{mappings[1]}{i}"].value,
            "min": sheet[f"{mappings[2]}{i}"].value,
            "max": sheet[f"{mappings[3]}{i}"].value,
            "on_hand": sheet[f"{mappings[4]}{i}"].value,
            "updated_by": session["user_id"],
        }
        this_warehouse = Warehouse.find_exact_by_code(data)
        if this_warehouse:
            warehouse_id = this_warehouse.id
        else:
            warehouse_id = 0
        this_item = Item.find_by_itemnumber_exact(data)
        if this_item:
            data["item_id"] = this_item.id
        else:
            data["item_id"] = 0
        data["warehouse_id"] = warehouse_id
        for key in data:
            if data[key] is None:
                if key == "min" or key == "max" or key == "on_hand":
                    data[key] = 0
                else:
                    data[key] = ""
            if key == "min" or key == "max" or key == "on_hand":
                data[key] = int(data[key])


        if request.form.get("update") == "on":
            result = Planning.validate_planning(data, imported=True, update=True)
        else:
            result = Planning.validate_planning(data, imported=True)
        if result[0]:
            flash(f"Successfully updated min/max for {data['item_number']} "
                      f"in {data['code']}", "import_successful")
            updated = Planning.update_from_whs_item(data)
            if not updated:
                Planning.save_planning(data)
                Quantity.save_quantity(data)
            Quantity.update_by_warehouse_item(data)
        else:
            message = f"Import failed for {data['item_number']} in {data['code']}: "
            for error in result[1]:
                message += error + " "
            flash(message, "import_fail")
    return redirect("/import/result")


@app.route("/import/quantities", methods=["POST"])
@logged_in
def import_counts():
    mappings = [request.form["count_item"],
                request.form["count_warehouse"],
                request.form["item_qty"]]

    file = request.files["item_count"]
    excel = load_workbook(file)
    sheet = excel.active

    if request.form["ignore_first_row"] == "on":
        start = 2
    else:
        start = 1

    end = sheet.max_row + 1

    for i in range(start, end):
        data = {
            "code": sheet[f"{mappings[1]}{i}"].value,
            "item_number": sheet[f"{mappings[0]}{i}"].value,
            "on_hand": sheet[f"{mappings[2]}{i}"].value,
            "updated_by": session.get("user_id")
        }
        for key in data:
            if data[key] is None:
                data[key] = ""
        result = Quantity.update_by_warehouse_item(data)

        if result:
            flash(
                f"Updated count for {data['item_number']} in "
                f"{data['code']} to {data['on_hand']}",
                "import_successful")
        else:
            flash(
                f"Could not update count for {data['item_number']} "
                f"in {data['code']}",
                "import_fail"
            )

    # ### TESTING quantity methods
    # data = {"code": mappings[1],
    #         "item_number": mappings[0],
    #         "on_hand": mappings[2],
    #         "updated_by": session.get("user_id")}
    # result = Quantity.update_by_warehouse_item(data)
    # flash(result, "import_successful")
    # #### END TEST ####

    return redirect("/import/result")


@app.route("/import/result")
@logged_in
def show_import_result():
    return render_template("import_finish.html")

from flask import Flask, render_template, session, flash, redirect, request, send_file
from flask_app import app
from flask_bcrypt import Bcrypt
from flask_app.models.warehouse import Warehouse
from flask_app.models.planning import Planning
from flask_app.models.item import Item
from flask_app.models.quantity import Quantity
from flask_app.controllers.users import logged_in
from openpyxl import Workbook, worksheet, load_workbook, writer
import os
from datetime import datetime

TODAY = datetime.today().strftime("%m-%d-%Y")

bcrypt = Bcrypt(app)

@app.route("/newWarehouse", methods=["POST"])
def add_warehouse():
    if(session["logged_in"]):
        data ={}
        for key in request.form:
            data[key] = request.form[key]
        if Warehouse.validate_warehouse(data):
            data["updated_by"] = session["user_id"]
            warehouse_id = Warehouse.save_warehouse(data)
            return redirect("/success")
        else:
            return redirect("/success")
    else:
        return redirect("/")

# @app.route("/warehouses/show/<int:id>")
# def show_warehouse(id):
#     if(session["logged_in"]):
#         data = {'id': id}
#         warehouse = Warehouse.find_by_id(data)
#         return render_template("showWarehouse.html", warehouse=warehouse)
#     else:
#         return redirect("/")

@app.route("/warehouses/show/<int:id>")
@logged_in
def show_warehouse(id):
    data = {'id': id}
    warehouse = Warehouse.find_by_id(data)
    if not warehouse:
        warehouse = Warehouse.empty_by_id(data)
    return render_template("showWarehouse.html", warehouse=warehouse)


@app.route("/warehouses/<int:id>/addPlanning", methods=["POST"])
@logged_in
def add_to_warehouse(id):
    data = {}
    for key in request.form:
        data[key] = request.form[key]
    data["warehouse_id"] = id
    data["updated_by"] = session["user_id"]
    if request.form["quantity"] == "" or request.form["quantity"] is None:
        data["on_hand"] = 0
    else:
        data["on_hand"] = int(request.form["quantity"])
    if Item.find_by_itemnumber_exact(request.form):
        data["item_id"] = Item.find_by_itemnumber_exact(request.form).id
    if Planning.validate_planning(data)[0]:
        Planning.save_planning(data)
        Quantity.save_quantity(data)
    return redirect(f"/warehouses/show/{id}")



# @app.route("/warehouses/<int:id>/addPlanning", methods=["POST"])
# def add_to_warehouse(id):
#     if session["logged_in"]:
#         data = {}
#         for key in request.form:
#             data[key] = request.form[key]
#         data["warehouse_id"] = id
#         data["updated_by"] = session["user_id"]
#         if request.form["quantity"] == "" or request.form["quantity"] is None:
#             data["on_hand"] = 0
#         else:
#             data["on_hand"] = int(request.form["quantity"])
#         if Item.find_by_itemnumber_exact(request.form):
#             data["item_id"] = Item.find_by_itemnumber_exact(request.form).id
#         if Planning.validate_planning(data):
#             Planning.save_planning(data)
#             Quantity.save_quantity(data)
#
#             return redirect(f"/warehouses/show/{id}")
#         else:
#             return redirect(f"/warehouses/show/{id}")
#     else:
#         return redirect("/")

@app.route("/warehouses/plannings/<int:planning_id>", methods=["GET", "POST"])
@logged_in
def update_planning(planning_id):
    if request.method == "GET":
        data = {'id': planning_id} #id corresponds to planning
        display_objects = Planning.find_planning_by_whs(data)
        display_planning = display_objects[0]
        display_item = display_objects[1]
        display_warehouse = display_objects[2]
        return render_template("planning.html",
                           display_planning=display_planning,
                           display_item=display_item,
                           display_warehouse=display_warehouse)
    else:
        warehouse_id = request.form["warehouse_id"]
        data = {
                "id": planning_id,
                "min": request.form["min"],
                "max": request.form["max"],
                "updated_by": session["user_id"]
            }
        if Planning.validate_update(data):
            Planning.update_planning(data)
            return redirect(f"/warehouses/show/{warehouse_id}")
        else:
            return redirect(f"/warehouses/plannings/{planning_id}")

@app.route("/warehouses/quantities/<int:qty_id>", methods=["POST"])
@logged_in
def update_quantity(qty_id):
    print("logged in!")
    data = {
            "id": qty_id,
            "on_hand": request.form["quantity"],
            "updated_by": session["user_id"]
        }
    warehouse = request.form["warehouse_id"]
    Quantity.update_quantity(data)
    return redirect(f"/warehouses/show/{warehouse}")

@app.route("/warehouses/download/<int:id>")
@logged_in
def download_warehouse_items(id):
    data = {'id': id}
    wb = Workbook()
    ws = wb.active

    ws["A1"].value = "Warehouse Report:"

    ws["A2"].value = TODAY
    #Table column names
    ws["A3"].value = "Item Number"
    ws["B3"].value = "Item Description"
    ws["C3"].value = "MIN"
    ws["D3"].value = "MAX"
    ws["E3"].value = "ON-HAND"
    ws["F3"].value = "Below Min?"
    ws["G3"].value = "Restock Quantity"

    row = 4


    warehouse = Warehouse.find_by_id(data)
    if warehouse:
        for line in warehouse.warehouse_items:
            ws[f"A{row}"].value = line.item_number
            ws[f"B{row}"].value = line.description
            ws[f"C{row}"].value = line.whs_min
            ws[f"D{row}"].value = line.whs_max
            ws[f"E{row}"].value = line.whs_qty
            if line.whs_qty < line.whs_min:
                ws[f"F{row}"].value = "YES"
            elif line.whs_qty < line.whs_max:
                ws[f"F{row}"].value = "Below Max"
            else:
                ws[f"F{row}"].value = "Fully Stocked"
            if line.whs_qty < line.whs_max:
                ws[f"G{row}"].value = line.whs_max - line.whs_qty
            else:
                ws[f"G{row}"].value = 0
            row += 1
    else:
        warehouse = Warehouse.empty_by_id(data)

    ws["B1"].value = warehouse.code + " " + warehouse.warehouse_name

    wb.save("flask_app/temp_files/TEMP.xlsx")


    # return "hello!"
    return send_file("temp_files/TEMP.xlsx", as_attachment=True, mimetype='application/vnd.ms-excel', attachment_filename=TODAY+warehouse.code+"export.xlsx")


